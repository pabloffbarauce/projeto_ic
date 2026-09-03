import numpy as np
from numpy import trapezoid
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def boltzmann_model(x, A1, A2, x0, dx):
    return A2 + (A1 - A2) / (1 + np.exp((x - x0) / dx))


def estimar_chute_inicial(x_data, y_data):
    x_data = np.asarray(x_data, dtype=float)
    y_data = np.asarray(y_data, dtype=float)

    A1_guess = y_data[0]
    A2_guess = y_data[-1]

    y_mid = A1_guess + (A2_guess - A1_guess) / 2
    idx = np.argmin(np.abs(y_data - y_mid))
    x0_guess = x_data[idx]

    dx_guess = (x_data.max() - x_data.min()) / 20.0

    return A1_guess, A2_guess, x0_guess, dx_guess


def boltzmann(x_data, y_data):
    x_data = np.asarray(x_data, dtype=float)
    y_data = np.asarray(y_data, dtype=float)

    p0 = estimar_chute_inicial(x_data, y_data)

    params, cov = curve_fit(
        boltzmann_model, x_data, y_data, p0=p0, method="lm", maxfev=10000
    )
    A1, A2, x0, dx = params

    y_fit = boltzmann_model(x_data, *params)
    treated_x_y_curve = list(zip(x_data, y_fit))

    return treated_x_y_curve, params, cov


def f_curve_calc(treated_x_y_curve):
    y_treated_values = [point[1] for point in treated_x_y_curve]
    y_treated_max = max(y_treated_values)

    f_curve = [y / y_treated_max for _, y in treated_x_y_curve]
    return f_curve, y_treated_max


def final_stats(x_data, f_curve):
    x_values_np = np.array(x_data)
    f_values_np = np.array(f_curve)

    E_t_curve = np.gradient(f_values_np, x_values_np)

    integrand = x_values_np * E_t_curve
    hydraulic_time = trapezoid(integrand, x_values_np)

    variance_calc = (x_values_np - hydraulic_time) ** 2 * E_t_curve
    variance = trapezoid(variance_calc, x_values_np)

    sigma_theta = variance / (hydraulic_time ** 2)

    N = 1 / sigma_theta if sigma_theta != 0 else 0

    return {
        "x_values_np": x_values_np,
        "E_t_curve": E_t_curve,
        "hydraulic_time": hydraulic_time,
        "variance": variance,
        "sigma_theta": sigma_theta,
        "N": N,
    }


def export_to_excel(path, x_data, y_data, treated_curve, f_curve, stats):
    y_adjusted = [p[1] for p in treated_curve]
    e_curve = stats["E_t_curve"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    headers = ["Tempo", "Concentração", "Concentração ajustada", "Curva F", "Curva E"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i in range(len(x_data)):
        ws.cell(row=i + 2, column=1, value=x_data[i])
        ws.cell(row=i + 2, column=2, value=y_data[i])
        ws.cell(row=i + 2, column=3, value=y_adjusted[i])
        ws.cell(row=i + 2, column=4, value=f_curve[i])
        ws.cell(row=i + 2, column=5, value=float(e_curve[i]))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    summary_labels = ["θh (tempo hidráulico médio)", "σ² (variância)",
                      "σθ² (variância adimensional)", "N (reatores em série)"]
    summary_values = [stats["hydraulic_time"], stats["variance"],
                      stats["sigma_theta"], stats["N"]]

    ws.cell(row=1, column=7, value="Resumo").font = header_font
    ws.cell(row=1, column=7).fill = header_fill
    ws.cell(row=1, column=8).fill = header_fill

    for i, (label, value) in enumerate(zip(summary_labels, summary_values), start=2):
        ws.cell(row=i, column=7, value=label)
        ws.cell(row=i, column=8, value=value)
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 16

    wb.save(path)
